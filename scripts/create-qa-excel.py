from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()
ws = wb.active
ws.title = "QA 테스트 시나리오"

hf = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
hfill = PatternFill("solid", fgColor="1B2A4A")
sf = Font(name="맑은 고딕", bold=True, size=11, color="1B2A4A")
sfill = PatternFill("solid", fgColor="E8E8E8")
bf = Font(name="맑은 고딕", size=10)
pfill = PatternFill("solid", fgColor="C6EFCE")
ffill = PatternFill("solid", fgColor="FFC7CE")
tb = Border(left=Side("thin", "D9D9D9"), right=Side("thin", "D9D9D9"), top=Side("thin", "D9D9D9"), bottom=Side("thin", "D9D9D9"))
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
la = Alignment(vertical="center", wrap_text=True)

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 6
ws.column_dimensions["D"].width = 48
ws.column_dimensions["E"].width = 48
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 25

ws.merge_cells("A1:G1")
ws["A1"] = "KECA 전체 기능 QA 테스트 시나리오"
ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14, color="1B2A4A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:G2")
ws["A2"] = "테스트 대상: https://keca.vercel.app  |  관리자: admin@keca.or.kr / 1234qwer  |  작성일: 2026-03-28"
ws["A2"].font = Font(name="맑은 고딕", size=9, color="555555")
ws["A2"].alignment = Alignment(horizontal="center")

for col, h in enumerate(["TC ID", "시나리오", "단계", "테스트 내용", "예상 결과", "결과", "비고/버그"], 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = hf; c.fill = hfill; c.alignment = ca; c.border = tb

S = "SEC"
tests = [
    (S,"1. 인증 시스템"),
    ("TC-1.1","회원가입","1","/register 접속","회원가입 폼 표시"),
    ("","","2","이름/이메일/연락처/비밀번호 입력","자동 하이픈 포맷팅 (연락처)"),
    ("","","3","비밀번호 6자 입력 후 제출","8자 이상 에러 메시지"),
    ("","","4","비밀번호 불일치 입력 후 제출","일치하지 않습니다 에러"),
    ("","","5","약관 미동의 상태에서 제출","버튼 비활성화"),
    ("","","6","이용약관/개인정보처리방침 클릭","새 탭에서 /terms, /privacy 열림"),
    ("","","7","정상 입력 후 제출","이메일 인증 메일 발송 성공 화면"),
    ("TC-1.2","로그인","1","/login 접속","로그인 폼 표시"),
    ("","","2","잘못된 비밀번호 입력","에러 메시지 표시"),
    ("","","3","관리자 계정으로 로그인","/mypage로 리다이렉트"),
    ("","","4","헤더 확인","관리자님 + 드롭다운 표시"),
    ("","","5","드롭다운에서 관리자 클릭","/admin 이동"),
    ("TC-1.3","비밀번호 재설정","1","비밀번호를 잊으셨나요? 클릭","/forgot-password 이동"),
    ("","","2","이메일 입력 후 재설정 링크 발송","이메일 발송 확인 화면"),
    ("TC-1.4","보호 라우트","1","비로그인 /mypage 접속","/login?redirect=/mypage"),
    ("","","2","비로그인 /admin 접속","/login?redirect=/admin"),
    ("","","3","일반 회원으로 /admin 접속","/mypage 리다이렉트"),
    ("TC-1.5","로그아웃","1","헤더 → 로그아웃 클릭","홈으로 리다이렉트, 헤더 복구"),
    (S,"2. 교육프로그램 관리"),
    ("TC-2.1","프로그램 등록","1","/admin/programs 접속","프로그램 목록 표시"),
    ("","","2","프로그램 등록 클릭","/admin/programs/new 폼"),
    ("","","3","프로그램명/카테고리/유형 입력","정상 입력"),
    ("","","4","썸네일 이미지 업로드","이미지 미리보기 표시"),
    ("","","5","등록 클릭","프로그램 목록으로 리다이렉트"),
    ("","","6","/programs 접속","새 프로그램 카드 표시"),
    ("","","7","상태를 공개로 변경 + 저장","저장 성공 메시지"),
    ("TC-2.2","프로그램 수정","1","편집 아이콘 클릭","수정 페이지 이동"),
    ("","","2","제목 수정 + 추천 프로그램 체크","정상 입력"),
    ("","","3","저장 클릭","프로그램이 저장되었습니다"),
    ("","","4","/programs/[slug] 접속","수정된 제목 반영"),
    ("TC-2.3","프로그램 삭제","1","삭제 클릭","confirm 다이얼로그"),
    ("","","2","확인 클릭","목록으로 리다이렉트"),
    ("TC-2.4","프로그램 상세","1","/programs/[slug] 접속","상세 정보 표시"),
    ("","","2","커리큘럼/기대효과 확인","모듈별 표시"),
    ("","","3","교육 문의하기 버튼 클릭","/inquiry 이동 (흰색 버튼)"),
    ("TC-2.5","검색/필터","1","검색창에 AI 입력","AI 프로그램만 필터링"),
    ("","","2","카테고리 탭 클릭","해당 카테고리만 표시"),
    (S,"3. 전문가과정"),
    ("TC-3.1","전문가과정 목록","1","/expert 접속","3개 그룹 표시"),
    ("","","2","과정 카드 클릭","/programs/[slug] 이동"),
    ("","","3","Admin 새 전문가과정 등록","등록 성공"),
    (S,"4. 강사진 관리"),
    ("TC-4.1","강사 등록","1","/admin/instructors 접속","강사 목록"),
    ("","","2","강사 등록 클릭","등록 폼"),
    ("","","3","프로필 이미지 업로드","미리보기"),
    ("","","4","프로필 카드 이미지 업로드","미리보기"),
    ("","","5","이름/이메일/전문분야/경력 입력","정상 입력"),
    ("","","6","강사 등록 클릭","목록으로 리다이렉트"),
    ("","","7","/instructors 접속","프로필 이미지 포함 카드"),
    ("","","8","카드 클릭 → 상세","프로필+카드이미지+YouTube"),
    ("TC-4.2","강사 수정","1","편집 아이콘 클릭","수정 페이지"),
    ("","","2","이미지 변경 + 저장","저장 성공"),
    ("TC-4.3","강사 삭제","1","삭제 → 확인","목록에서 제거"),
    (S,"5. 커뮤니티 게시판"),
    ("TC-5.1","공지사항","1","/admin/community/notice","관리 페이지"),
    ("","","2","글 작성 → 등록","등록 성공"),
    ("","","3","/community/notice 확인","새 공지 (핀 아이콘)"),
    ("","","4","공지 클릭 → 상세","제목/내용/날짜/조회수"),
    ("","","5","수정 → 저장","수정 성공"),
    ("","","6","삭제 → 확인","삭제 성공"),
    ("","","7","검색창 키워드 입력","필터링 결과"),
    ("TC-5.2","사진갤러리","1","/admin/community/photo_gallery","갤러리 관리"),
    ("","","2","글 + 이미지 업로드 → 등록","업로드+등록 성공"),
    ("","","3","/community/gallery 확인","이미지 카드 (hover scale)"),
    ("TC-5.3","영상갤러리","1","/admin/community/video_gallery","영상 관리"),
    ("","","2","글 + YouTube URL → 등록","등록 성공"),
    ("","","3","/community/videos 확인","YouTube 썸네일"),
    ("","","4","카드 → 상세","iframe 임베드 재생"),
    ("TC-5.4","자료실","1","/admin/community/resource","자료실 관리"),
    ("","","2","글 + 파일 업로드 → 등록","업로드+등록 성공"),
    ("","","3","/community/resources","다운로드 아이콘 활성화"),
    ("","","4","다운로드 클릭","파일 다운로드"),
    ("TC-5.5","교육후기","1","Admin 후기 글 작성","등록 성공"),
    ("","","2","/community/review 확인","후기 카드 (별점)"),
    ("","","3","홈페이지 교육후기 섹션","최신 3건 표시"),
    (S,"6. 교육일정"),
    ("TC-6.1","일정 CRUD","1","/admin/schedules","일정 관리"),
    ("","","2","일정 추가 → 저장","등록 성공"),
    ("","","3","/community/schedule","날짜 배지 + 카드"),
    ("","","4","홈페이지 교육일정","미래 일정 3건"),
    ("","","5","Admin 삭제","삭제 + 공개 반영"),
    (S,"7. 교육문의"),
    ("TC-7.1","문의 접수→관리","1","/inquiry 접속","교육문의 폼"),
    ("","","2","필수 필드 입력 후 제출 (하늘색)","접수 성공 화면"),
    ("","","3","/admin/inquiries","문의 표시 (신규)"),
    ("","","4","상태 처리중 → 완료 변경","상태 변경 성공"),
    (S,"8. 카테고리 관리"),
    ("TC-8.1","카테고리 CRUD","1","/admin/categories","카테고리 그리드"),
    ("","","2","분야 추가 → 이름/slug → 추가","새 카테고리 카드"),
    ("","","3","/programs 카테고리 탭 확인","새 탭 표시"),
    ("","","4","편집 → 수정","수정 반영"),
    ("","","5","삭제 → 확인","삭제 반영"),
    (S,"9. 협회/자격소개 관리"),
    ("TC-9.1","협회소개 관리","1","/admin/about","7개 탭"),
    ("","","2","인사말 수정 → 전체 저장","저장 성공"),
    ("","","3","/about 확인","수정 내용 반영"),
    ("TC-9.2","자격소개 관리","1","/admin/certification","4개 탭"),
    ("","","2","수정 → 저장","저장 성공"),
    ("","","3","/certification 확인","수정 반영"),
    (S,"10. 회원관리"),
    ("TC-10.1","회원 승인/역할","1","/admin/members","회원 목록"),
    ("","","2","승인대기 필터","pending 회원"),
    ("","","3","승인 버튼 클릭","승인 성공 + 이메일"),
    ("","","4","역할 정회원 변경","역할 변경 성공"),
    ("","","5","이름/이메일 검색","필터링 결과"),
    (S,"11. 파트너 관리"),
    ("TC-11.1","파트너 CRUD","1","/admin/partners","파트너 목록"),
    ("","","2","추가 → 기관명/웹사이트 → 저장","추가 성공"),
    ("","","3","홈페이지 파트너 섹션","파트너명 표시"),
    ("","","4","삭제 → 홈페이지 확인","섹션에서 제거"),
    (S,"12. 사이트 설정"),
    ("TC-12.1","설정 저장","1","/admin/settings","설정 폼"),
    ("","","2","이메일/협회비/SNS 수정 → 저장","저장 성공"),
    (S,"13. 결제 시스템"),
    ("TC-13.1","협회비 결제","1","회원 → /mypage/membership","납부 상태"),
    ("","","2","납부하기 → /pay","결제 카드 (100,000원)"),
    ("","","3","결제하기 클릭","토스페이먼츠 결제창"),
    ("","","4","결제 완료","성공 페이지"),
    ("","","5","/admin/payments","결제 내역"),
    (S,"14. 홈페이지 동적 데이터"),
    ("TC-14.1","홈페이지 섹션","1","/ 접속","풀스크린 히어로 + SCROLL"),
    ("","","2","헤더 메뉴 색상","홈:흰색, 스크롤:다크"),
    ("","","3","카테고리 위젯","8개 아이콘+링크"),
    ("","","4","실적 카운터","500+/50+/30+/98%"),
    ("","","5","추천 교육","featured 최대 3개"),
    ("","","6","공지사항","DB 최신 4건"),
    ("","","7","교육문의 바로가기","흰색 pill → /inquiry"),
    ("","","8","교육후기","DB 최신 3건"),
    ("","","9","교육일정","미래 일정 3건"),
    ("","","10","파트너 섹션","DB 파트너"),
    ("","","11","CTA 섹션","회원가입+회원혜택 버튼"),
    (S,"15. UI/UX 공통"),
    ("TC-15.1","반응형 디자인","1","모바일(375px) 홈페이지","햄버거 메뉴"),
    ("","","2","모바일 메뉴 열기","드로어 + 인증"),
    ("","","3","태블릿(768px) 프로그램","2열 카드"),
    ("","","4","데스크톱(1280px)","3열 카드+전체 네비"),
    ("TC-15.2","접근성","1","Tab → 본문 건너뛰기","포커스 시 표시"),
    ("","","2","aria-label 확인","메인 메뉴"),
    ("TC-15.3","에러 처리","1","존재하지 않는 URL","커스텀 404"),
    ("","","2","콘솔 에러 확인","에러 0건"),
    (S,"16. 정적 페이지"),
    ("TC-16.1","정적 페이지","1","/terms 접속","이용약관 6개 조항"),
    ("","","2","/privacy 접속","개인정보 7개 항목"),
    ("","","3","/about ~ /about/location","7개 모두 200 OK"),
    ("","","4","/certification ~ /graduates","4개 모두 200 OK"),
]

row = 4
for t in tests:
    if t[0] == S:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws.cell(row=row, column=1, value=t[1])
        c.font = sf; c.fill = sfill; c.alignment = la; c.border = tb
        ws.row_dimensions[row].height = 25
    else:
        vals = list(t) + [""] * (7 - len(t))
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = bf; c.border = tb
            c.alignment = ca if col in (3, 6) else la
        ws.row_dimensions[row].height = 30
    row += 1

dv = DataValidation(type="list", formula1='"PASS,FAIL,SKIP,N/A"', allow_blank=True)
ws.add_data_validation(dv)
for r in range(4, row):
    if ws.cell(row=r, column=1).value not in (None, S) or ws.cell(row=r, column=3).value:
        dv.add(ws.cell(row=r, column=6))

ws.conditional_formatting.add(f"F4:F{row-1}", CellIsRule(operator="equal", formula=['"PASS"'], fill=pfill))
ws.conditional_formatting.add(f"F4:F{row-1}", CellIsRule(operator="equal", formula=['"FAIL"'], fill=ffill))

ws.freeze_panes = "A4"
ws.page_setup.orientation = "landscape"

wb.save("docs/KECA-QA-테스트시나리오.xlsx")
print(f"Created: docs/KECA-QA-테스트시나리오.xlsx ({row-4} rows)")
